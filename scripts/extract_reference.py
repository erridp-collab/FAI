"""Estrae ogni foglio di FAI_Microimpresa_v6(1).xlsx in un markdown leggibile.
Cattura valori e formule (data_only=False) per documentare la fonte di verità.
Uso (dalla root del progetto, cioè la cartella FAI padre di fai-web):
    ./venv/Scripts/python.exe scripts/extract_reference.py
Oppure dalla root fai-web/ con percorsi relativi aggiornati:
    ../venv/Scripts/python.exe scripts/extract_reference.py
"""
import os
import openpyxl

SRC = "FAI_Microimpresa_v6(1).xlsx"
OUT = "fai-web/docs/product-readiness/REFERENCE_RAW.md"

def cell_repr(cell):
    val = cell.value
    if val is None:
        return None
    if isinstance(val, str) and val.startswith("="):
        return f"FORMULA `{val}`"
    return str(val).strip()

def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = openpyxl.load_workbook(SRC, data_only=False)
    lines = ["# REFERENCE_RAW — dump grezzo Excel", ""]
    lines.append(f"Generato da `scripts/extract_reference.py` su `{SRC}`.")
    lines.append("Non modificare a mano: rigenerabile. La curatela vive in REFERENCE.md.")
    lines.append("")
    for name in wb.sheetnames:
        ws = wb[name]
        lines.append(f"## Foglio: {name}")
        lines.append("")
        lines.append("| Cella | Contenuto |")
        lines.append("|---|---|")
        for row in ws.iter_rows():
            for cell in row:
                rep = cell_repr(cell)
                if rep is not None:
                    safe = rep.replace("|", "\\|").replace("\n", " ")
                    lines.append(f"| {cell.coordinate} | {safe} |")
        lines.append("")
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"Scritto {OUT} ({len(lines)} righe)")

if __name__ == "__main__":
    main()
